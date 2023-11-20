import json
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from entities.colloquium import Colloquium
from entities.speaker import Speaker


# Substitua 'caminho_do_arquivo_json' pelo caminho do seu arquivo JSON
caminho_do_arquivo_json = 'C:\\Users\\junio\\Documents\\certificados\\datas\\datas.json'

# Ler o arquivo JSON
with open(caminho_do_arquivo_json, 'r', encoding='utf-8') as file:
    data = json.load(file)

# Imprimir o evento de cada entrada no JSON
colloquiums = []
for entry in data:
    colloquiums.append(
        Colloquium(
            date=entry['date'],
            name=entry['event'],
            speakers = [Speaker(**speaker) for speaker in entry['speakers']]
        )
    )

# Define the columns for the spreadsheet
columns = [
    "Nome",
    "Data Nasc.",
    "Tipo Doc.",
    "Número do Documento",
    "Endereço",
    "Email",
    "Nome da Mãe",
    "Tipo de Participação",
    "CH",
    "Extra 1",
    "Extra 2",
    "Extra 3"
]

# Columns to be highlighted
highlighted_columns = [
    "Nome",
    "Data Nasc.",
    "Tipo Doc.",
    "Número do Documento",
    "Tipo de Participação",
    "CH"
]


for colloquium in colloquiums:
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the styles for header cells
    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Apply bold font and red background to specified columns
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = bold_font
        if column_title in highlighted_columns:
            cell.fill = red_fill


    # Adiciona os dados nas linhas correspondentes às colunas destacadas
    for row, speaker in enumerate(colloquium.speakers, 2):
        for col, column in enumerate(columns, 1):
            cell = ws.cell(row=row, column = col)
            
            if column == "Nome":
                cell.value = speaker.name
            elif column == "Data Nasc.":
                cell.value = speaker.birth_date
            elif column == "Tipo Doc.":
                cell.value = "CPF"
            elif column == "Número do Documento":
                cell.value = speaker.cpf
            elif column == "Tipo de Participação":
                cell.value = "Palestrante"
            elif column == "CH":
                cell.value = "1"               

    # Create a directory 'pasta_sub' inside the current directory if it doesn't exist
    dir_path = 'excels'
    os.makedirs(dir_path, exist_ok=True)

    # Save the workbook to the 'pasta_sub' directory with the given filename
    file_path = os.path.join(dir_path, f'{colloquium.get_name_format()}.xlsx')
    wb.save(file_path)

    # Print the path where the file is saved (for confirmation purposes)
    print(f"Planilha salva em: {file_path}")
